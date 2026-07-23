from typing import Any, List
import asyncpg
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

async def get_latest_calculations(conn: asyncpg.Connection, limit: int = 20) -> List[dict[str, Any]]:
    rows = await conn.fetch("""
        SELECT id, name, date1 as calculated_at, fileid
        FROM calculation
        ORDER BY date1 DESC NULLS LAST, id DESC
        LIMIT $1
    """, limit)
    return [dict(row) for row in rows]

async def get_calculation_results_geojson(conn: asyncpg.Connection, calculation_id: int) -> dict[str, Any]:
    rows = await conn.fetch("""
        SELECT
            line.id as line_id,
            line.name as label,
            ut.a7 as flow,
            ut.a8 as velocity,
            ut.a9 as pressure_drop,
            ut.a10 as specific_pressure_drop,
            ut.a11 as head_start,
            ut.a12 as head_end,
            ST_AsGeoJSON(ST_Transform(line.shape, 4326)) AS geometry
        FROM ut_out ut
        JOIN linesobj line ON line.id = ut.lineid
        WHERE ut.calculationid =  AND line.shape IS NOT NULL
    """, calculation_id)
    
    import json
    features = []
    for row in rows:
        geom = row["geometry"]
        if not geom:
            continue
            
        features.append({
            "type": "Feature",
            "geometry": json.loads(geom),
            "properties": {
                "id": row["line_id"],
                "label": row["label"] or "",
                "flow": row["flow"],
                "velocity": row["velocity"],
                "pressure_drop": row["pressure_drop"],
                "specific_pressure_drop": row["specific_pressure_drop"],
                "head_start": row["head_start"],
                "head_end": row["head_end"]
            }
        })
        
    return {
        "type": "FeatureCollection",
        "features": features
    }

async def get_calculation_results_excel(conn: asyncpg.Connection, calculation_id: int) -> bytes:
    rows = await conn.fetch("""
        SELECT
            line.id as line_id,
            line.name as label,
            ut.a7 as flow,
            ut.a8 as velocity,
            ut.a9 as pressure_drop,
            ut.a10 as specific_pressure_drop,
            ut.a11 as head_start,
            ut.a12 as head_end
        FROM ut_out ut
        JOIN linesobj line ON line.id = ut.lineid
        WHERE ut.calculationid = 
        ORDER BY line.id ASC
    """, calculation_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Расчет #{calculation_id}"

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    headers = [
        "ID участка", "Наименование", "Расход (т/ч)", "Скорость (м/с)",
        "Падение давления (атм)", "Удельное падение (мм/м)",
        "Напор в начале (м)", "Напор в конце (м)"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r['line_id'], r['label'] or '', r['flow'], r['velocity'],
            r['pressure_drop'], r['specific_pressure_drop'],
            r['head_start'], r['head_end']
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
