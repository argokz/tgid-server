"""Генерация печатных форм (HTML) и ведомостей (Excel).

Данные берутся из тех же проверенных read-only функций, что и журналы web-UI,
поэтому имена полей всегда соответствуют реальной схеме БД. Прямой SQL здесь
раньше ссылался на несуществующие колонки (nodes.name, networkarmatures) и
падал с 500 — теперь источник данных один на журнал и на отчёт.
"""

import io
import os
from typing import Any, Callable, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database.connect import acquire_conn
from database.consumer_load_diagnostics import get_consumer_load_diagnostics
from database.defects import get_defects
from database.inspections import get_inspections
from database.network_armatures import get_network_armatures
from database.network_bypasses import get_network_bypasses
from database.pressure_tests import get_pressure_tests
from database.pump_equipment import get_installed_pumps
from database.repairs import get_repairs
from database.shurfs import get_shurfs

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "report_templates")

MAX_REPORT_ROWS = 5000


def get_template_content(filename: str) -> str:
    path = os.path.join(TEMPLATE_DIR, filename)
    if os.path.exists(path):
        with open(path, "r", encoding="windows-1251", errors="ignore") as f:
            return f.read()
    return ""


def _esc(value: Any) -> str:
    """Экранирование значения для вставки в HTML-таблицу."""
    if value is None:
        return ""
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _object_title(item: Dict[str, Any]) -> str:
    if item.get("line_id"):
        return f"Участок #{item['line_id']}"
    if item.get("node_id"):
        return f"Узел #{item['node_id']}"
    return "—"


# --- HTML-формы -------------------------------------------------------------

FORM_TITLES = {
    "f10_remont": "Форма 10. Журнал ремонтов",
    "f11_defect": "Форма 11. Журнал нарушений",
    "f12_pits": "Форма 12. Журнал шурфовок",
    "f13_inspection": "Форма 13. Журнал осмотров",
    "f14_pressure": "Форма 14. Журнал опрессовок",
}

FORM_COLUMNS: Dict[str, List[str]] = {
    "f10_remont": ["Объект", "Название", "Вид ремонта", "Состояние", "Категория",
                   "План: начало", "План: окончание", "Ответственный"],
    "f11_defect": ["Объект", "Название", "Источник", "Состояние", "Категория",
                   "Дата обнаружения", "Описание"],
    "f12_pits": ["Объект", "Название", "Назначение", "Состояние", "Дата вскрытия", "Утверждение"],
    "f13_inspection": ["Название", "Дата", "Ответственный", "Подразделение", "Акт"],
    "f14_pressure": ["Название", "Состояние", "Вид испытания", "Источник тепла",
                     "План: начало", "Ответственный"],
}


def _form_rows(form_id: str, items: List[Dict[str, Any]]) -> List[List[Any]]:
    if form_id == "f10_remont":
        return [[
            _object_title(i), i.get("name"), i.get("repair_type_name"), i.get("state_name"),
            i.get("category_name"), i.get("planned_start"), i.get("planned_finish"),
            i.get("responsible_name"),
        ] for i in items]
    if form_id == "f11_defect":
        return [[
            _object_title(i), i.get("name"), i.get("source_name"), i.get("state_name"),
            i.get("category_name"), i.get("detected_at"), i.get("description"),
        ] for i in items]
    if form_id == "f12_pits":
        return [[
            _object_title(i), i.get("name"), i.get("purpose_name"), i.get("state_name"),
            i.get("opening_date") or i.get("detected_at"), i.get("approval_name"),
        ] for i in items]
    if form_id == "f13_inspection":
        return [[
            i.get("name"), i.get("inspection_date") or i.get("detected_at"),
            i.get("responsible_name"), i.get("subdivision_name"), i.get("act_number"),
        ] for i in items]
    if form_id == "f14_pressure":
        return [[
            i.get("name"), i.get("state_name"), i.get("test_type_name"),
            i.get("heat_source_name"), i.get("planned_start"), i.get("responsible_name"),
        ] for i in items]
    return []


FORM_LOADERS: Dict[str, Callable] = {
    "f10_remont": get_repairs,
    "f11_defect": get_defects,
    "f12_pits": get_shurfs,
    "f13_inspection": get_inspections,
    "f14_pressure": get_pressure_tests,
}


async def generate_form_html(form_id: str, search: Optional[str] = None) -> str:
    title = FORM_TITLES.get(form_id, f"Отчёт {form_id}")
    loader = FORM_LOADERS.get(form_id)

    if loader is None:
        return _render_html(title, [], [], note=f"Неизвестная форма отчёта: {form_id}")

    async with acquire_conn() as conn:
        data = await loader(conn, page=1, page_size=1000, search=search)

    items = data.get("items", []) if isinstance(data, dict) else []
    columns = FORM_COLUMNS.get(form_id, [])
    rows = _form_rows(form_id, items)

    note = None
    if not rows:
        note = "По заданным условиям записей не найдено."

    body = _render_html(title, columns, rows, note=note, total=data.get("total") if isinstance(data, dict) else None)

    # Если для формы есть legacy-шаблон с таблицей — вставляем строки в него
    base_html = get_template_content(f"{form_id}.html")
    if base_html and "</table>" in base_html:
        rows_html = "".join(
            "<tr>" + "".join(f"<td>{_esc(c)}</td>" for c in row) + "</tr>" for row in rows
        )
        head, _, tail = base_html.partition("</table>")
        return head + rows_html + "</table>" + tail

    return body


def _render_html(
    title: str,
    columns: List[str],
    rows: List[List[Any]],
    note: Optional[str] = None,
    total: Optional[int] = None,
) -> str:
    head_html = "".join(f"<th>{_esc(c)}</th>" for c in columns)
    rows_html = "".join(
        "<tr>" + "".join(f"<td>{_esc(c)}</td>" for c in row) + "</tr>" for row in rows
    )
    note_html = f'<p class="note">{_esc(note)}</p>' if note else ""
    total_html = f'<p class="meta">Всего записей: {total}</p>' if total is not None else ""
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <title>{_esc(title)}</title>
  <style>
    body {{ font-family: system-ui, sans-serif; margin: 24px; color: #1a1a1a; }}
    h2 {{ font-size: 18px; margin: 0 0 4px; }}
    .meta, .note {{ color: #555; font-size: 13px; margin: 4px 0 12px; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th, td {{ border: 1px solid #ccc; padding: 6px 8px; font-size: 13px; text-align: left; vertical-align: top; }}
    th {{ background: #1f497d; color: #fff; position: sticky; top: 0; }}
    tr:nth-child(even) td {{ background: #fafafa; }}
    @media print {{ body {{ margin: 0; }} th {{ position: static; }} }}
  </style>
</head>
<body>
  <h2>{_esc(title)}</h2>
  {total_html}
  {note_html}
  <table><thead><tr>{head_html}</tr></thead><tbody>{rows_html}</tbody></table>
</body>
</html>"""


# --- Excel-ведомости --------------------------------------------------------

async def _rows_pipelines(conn) -> tuple[List[str], List[List[Any]]]:
    headers = ["ID участка", "Узел 1", "Узел 2", "Длина, м", "Ø внутр., мм",
               "Ø условный, мм", "Ø наружн., мм", "Толщина стенки, мм"]
    rows = await conn.fetch(
        """
        SELECT L.id, L.nodeid1, L.nodeid2,
               HPS.pipesectlength, HPS.diameterinternal, HPS.diametercondit,
               HPS.diameterexternal, HPS.wallthickness
        FROM linesobj L
        LEFT JOIN heatpipesections HPS ON HPS.lineid = L.id
        WHERE COALESCE(L.removed, 0) = 0
        ORDER BY L.id
        LIMIT $1
        """,
        MAX_REPORT_ROWS,
    )
    return headers, [
        [r["id"], r["nodeid1"], r["nodeid2"], r["pipesectlength"], r["diameterinternal"],
         r["diametercondit"], r["diameterexternal"], r["wallthickness"]]
        for r in rows
    ]


async def _rows_armatures(conn) -> tuple[List[str], List[List[Any]]]:
    data = await get_network_armatures(conn, page=1, page_size=MAX_REPORT_ROWS)
    headers = ["ID", "Участок", "Наименование", "Назначение", "Ø условный, мм",
               "Состояние", "Открытие, %", "Число оборотов", "Типоразмер"]
    return headers, [
        [i.get("id"), i.get("line_id"), i.get("display_name"), i.get("purpose_name"),
         i.get("nominal_diameter"), i.get("state_name"), i.get("opening_percent"),
         i.get("turn_count"), i.get("standard_mark")]
        for i in data.get("items", [])
    ]


async def _rows_bypasses(conn) -> tuple[List[str], List[List[Any]]]:
    data = await get_network_bypasses(conn, page=1, page_size=MAX_REPORT_ROWS)
    headers = ["ID", "Участок", "Наименование", "Узел подключения", "Состояние",
               "Трубопровод", "Длина, м", "Ø внутр., мм", "Расход (задание)", "Напор (задание)"]
    return headers, [
        [i.get("id"), i.get("line_id"), i.get("display_name"), i.get("connection_node_id"),
         i.get("state_name"), i.get("pipeline_sign_name"), i.get("length"),
         i.get("internal_diameter"), i.get("set_flow"), i.get("set_head")]
        for i in data.get("items", [])
    ]


async def _rows_pumps(conn) -> tuple[List[str], List[List[Any]]]:
    data = await get_installed_pumps(conn, page=1, page_size=MAX_REPORT_ROWS)
    headers = ["ID", "Участок", "Номер", "Насосная станция", "Модель", "Тип",
               "Кол-во агрегатов", "Тип привода", "Состояние", "Фрагмент"]
    return headers, [
        [i.get("id"), i.get("line_id"), i.get("number"), i.get("station_name"),
         i.get("model_name"), i.get("model_type"), i.get("parallel_count"),
         i.get("drive_type_name"), i.get("state_name"), i.get("fragment_name")]
        for i in data.get("items", [])
    ]


async def _rows_consumers(conn) -> tuple[List[str], List[List[Any]]]:
    data = await get_consumer_load_diagnostics(conn, page=1, page_size=MAX_REPORT_ROWS)
    headers = ["Тип", "ID", "Узел", "Наименование", "Состояние", "Отопление, Гкал/ч",
               "Вентиляция, Гкал/ч", "ГВС, Гкал/ч", "Итого, Гкал/ч", "Фрагмент"]
    type_names = {"generalized": "Обобщённый", "real": "Реальный"}
    return headers, [
        [type_names.get(i.get("consumer_type"), i.get("consumer_type")), i.get("id"),
         i.get("node_id"), i.get("name"), i.get("state_name"), i.get("heating_load"),
         i.get("ventilation_load"), i.get("hot_water_load"), i.get("total_load"),
         i.get("fragment_name")]
        for i in data.get("items", [])
    ]


EXCEL_SHEETS: Dict[str, tuple[str, Callable]] = {
    "ut": ("Участки теплопроводов", _rows_pipelines),
    "pipelines": ("Участки теплопроводов", _rows_pipelines),
    "zd": ("Задвижки и арматура", _rows_armatures),
    "valves": ("Задвижки и арматура", _rows_armatures),
    "bp": ("Байпасы", _rows_bypasses),
    "bypasses": ("Байпасы", _rows_bypasses),
    "ns": ("Насосные агрегаты", _rows_pumps),
    "pumps": ("Насосные агрегаты", _rows_pumps),
    "pt": ("Потребители", _rows_consumers),
    "consumers": ("Потребители", _rows_consumers),
}


def excel_report_types() -> List[Dict[str, str]]:
    """Уникальные ведомости для UI: код + название листа."""
    seen: Dict[str, str] = {}
    for code, (title, _) in EXCEL_SHEETS.items():
        seen.setdefault(title, code)
    return [{"code": code, "title": title} for title, code in seen.items()]


async def generate_excel_report(doc_type: str) -> bytes:
    entry = EXCEL_SHEETS.get(doc_type.lower())
    if entry is None:
        raise ValueError(
            f"Неизвестный тип ведомости: {doc_type}. Доступны: {', '.join(sorted(EXCEL_SHEETS))}"
        )

    sheet_title, loader = entry
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    async with acquire_conn() as conn:
        headers, rows = await loader(conn)

    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ширина колонок по содержимому первых строк + фиксация шапки и автофильтр
    for idx, header in enumerate(headers, start=1):
        width = len(str(header))
        for row in rows[:200]:
            value = row[idx - 1] if idx - 1 < len(row) else None
            if value is not None:
                width = max(width, min(len(str(value)), 60))
        ws.column_dimensions[get_column_letter(idx)].width = width + 3
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
