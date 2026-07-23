from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import NamedStyle

import logging

import re
import time

import datetime
import psycopg2 as pyodbc
import hyp

#-------------------------------------------------------------------------------------

# Функция для расчета высоты строки
def calculate_row_height(text, font_size, column_width):
    lines = text.split("\n")
    num_lines = sum(len(line) // column_width + 1 for line in lines)
    return num_lines * (font_size * 0.75)  # примерное значение для высоты строки

#-------------------------------------------------------------------------------------

# Определение стиля границы (тонкая черная линия)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

#-------------------------------------------------------------------------------------

#thin_border = Border(
hair_border = Border(

    left=Side(style='hair', color='000000'),
    right=Side(style='hair', color='000000'),
    top=Side(style='hair', color='000000'),
    bottom=Side(style='hair', color='000000')
)

#-------------------------------------------------------------------------------------


underline_border = Border(
    bottom=Side(style='hair', color='000000')
)

#-------------------------------------------------------------------------------------


no_border = Border()


#-------------------------------------------------------------------------------------

custom_font = Font(name='Arial', size=14, bold=True, italic=False, color="FF0000")
#custom_font = Font(name='Arial', size=14, bold=True, italic=False, color="FF0000")
small_font = Font(size=8, bold=False, italic=False)

#-------------------------------------------------------------------------------------

small_alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
memo_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
top_alignment = Alignment(vertical='top', wrap_text=True)

#-------------------------------------------------------------------------------------

#date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY HH:MM:MM')
date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')

#-------------------------------------------------------------------------------------


#merged_cell.font = custom_font

#-------------------------------------------------------------------------------------

def write_text(sheet, row, col, value, border=no_border, alignment=None, bold=False):
    cell = sheet.cell(row=row, column=col)
    cell.value = value

    if isinstance(value, datetime.date):
        cell.style = date_style

#    if isinstance(cell.value, str):
#        cell.value = hyp.hyphen(cell.value)

    cell.border = border  # Применение границы к ячейке

    if alignment:
        cell.alignment = alignment

    cell.font = cell.font.copy(name="Times New Roman")

    if bold:
        cell.font = cell.font.copy(bold=True)


    return cell

#-------------------------------------------------------------------------------------


def write_text_cell(sheet, cell, value, border=no_border, alignment=None):
#    cell = sheet.cell(row=row, column=col)
    cell.value = value
    cell.border = border  # Применение границы к ячейке

    if isinstance(value, datetime.date):
        cell.style = date_style

    if alignment:
        cell.alignment = alignment

    return cell


#-------------------------------------------------------------------------------------

def a2n(cell):
    column_letter, row_number = coordinate_from_string(cell)
    column_number = column_index_from_string(column_letter)
    return row_number, column_number

#-------------------------------------------------------------------------------------

def a4n(cell_range):
    start_cell, end_cell = cell_range.split(':')
    return a2n(start_cell), a2n(end_cell) 

#-------------------------------------------------------------------------------------


def write_text2(sheet, cell_range, value, border=no_border, alignment=None, dx=0, dy=0, bold=False):
    if ':' in cell_range:
        (row1, col1), (row2, col2) = a4n(cell_range)

        if dx > 0 or dy > 0:
            row1, col1 = row1+dy, col1+dx
            row2, col2 = row2+dy, col2+dx
            cell_range = f'{get_column_letter(col1)}{row1}:{get_column_letter(col2)}{row2}'

        sheet.merge_cells(cell_range)
        cell = write_text(sheet, row1, col1, value, border=border, alignment=alignment)

        for row in range(row1, row2+1):
            for col in range(col1, col2+1):
                c = sheet.cell(row=row, column=col)
                c.border = border
                if alignment:
                    c.alignment = alignment

                if bold:
                    c.font = c.font.copy(bold=True)


        return cell

    else:
        (row1, col1) = a2n(cell_range)
        if dx > 0 or dy > 0:
            row1, col1 = row1+dy, col1+dx

#    return write_text_cell(sheet, cell, value, border=border, alignment=alignment)
        return write_text(sheet, row1, col1, value, border=border, alignment=alignment, bold=bold)

#-------------------------------------------------------------------------------------

def write_row_old(sheet, start_row, start_col, vals, border=no_border):
#    for col_index, column_name in enumerate(cursor.description):
#        sheet.cell(row=start_row, column=start_col + col_index, value=column_name[0])

    for col_index, val in enumerate(vals):
        alignment = top_alignment

#        if sheet.cell()
        
        if len(str(val)) > 45:
            alignment = memo_alignment
#            sheet.row_dimensions[start_row].height = rsgtdrt
            sheet.row_dimensions[start_row].bestFit  = True


        write_text(sheet, start_row, start_col + col_index, val, border, alignment=alignment)

#        cell = sheet.cell(row=start_row, column=start_col + col_index, value=val)
#        cell.border = thin_border  # Применение границы к ячейке
#        cell.font = custom_font

#-------------------------------------------------------------------------------------


def write_row(sheet, vals):
    row = []

    for col_index, val in enumerate(vals):
        row.append(val)

    sheet.append(row)            

#-------------------------------------------------------------------------------------


def write_table(ws, conn, q, row0=1, col0=1, numbers=True, freez=True):

    try:
        print(f'Начал {ws.title}')
        logging.info(f'Начал {ws.title}')
        
        t1 = time.time()

        cursor = conn.cursor()

        cursor.execute(q)

        columns = [column[0] for column in cursor.description]
       
        r, c = row0, col0


        if numbers:
            for c in range(len(columns)):
                write_text(ws, r, c+1, f'{c+1}', thin_border, alignment=center_alignment, bold=False)
            r += 1

        row_start = r

#        if freez:
#            ws.freeze_panes = ws[f'A{row_start}']
    
    #    print(f'Начал печатать {ws.title}')

        while True:
            row = cursor.fetchone()
            if not row: break
#            write_row_old(ws, r, col0, list(row), thin_border)
            write_row(ws, list(row))
            r += 1

        for row in ws.iter_rows(min_row=row_start, max_row=r-1, min_col=col0, max_col=col0+len(columns)-1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = top_alignment
#                cell.font = font
#                cell.fill = fill

        if freez:
            ws.freeze_panes = ws[f'A{row_start}']

        cursor.close()

        t2 = time.time()


        adjust_table(ws, row0, col0, r, len(columns))

        t3 = time.time()

        print('Закончил', round((t3-t1)*1000), round((t2-t1)*1000), round((t3-t2)*1000))
        logging.info(f'Закончил {ws.title} {round((t3-t1)*1000)} {round((t2-t1)*1000)} {round((t3-t2)*1000)}')


        return r-1, col0 + len(columns)-1

    except pyodbc.Error as e:
        
        logging.warning(f'Ошибка')
        logging.warning(e)
        logging.warning(f'-------------------------------------------')
        logging.warning(q)
        logging.warning(f'-------------------------------------------')
        print(f'-------------------------------------------')
        print(e)
#        print(f'-------------------------------------------')
        print(q)
        print(f'-------------------------------------------')
        exit(3)

    return r, c

#-------------------------------------------------------------------------------------

def adjust_table(ws, row1, col1, row2, col2):
    # Автоматически устанавливаем ширину столбцов

    max_max_len = 45
    for col_idx in range(col1, col2+1):  # Проходим только по тем столбцам, где есть данные
        max_length = 0
        column_letter = get_column_letter(col_idx)  # Получаем букву столбца
        for row_idx in range(row1, row2+1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            try:
                if len(str(cell_value)) > max_max_len:
                    max_length = max_max_len
                    break
                if len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            except:
                logging
                pass
        adjusted_width = max_length + 2  # Немного добавляем к длине
        ws.column_dimensions[column_letter].width = adjusted_width

    for col_idx in range(col1, col2+1):  # Проходим заголовкам
        max_length = 0
#        print('----------------------------', col_idx)
        
        column_letter = get_column_letter(col_idx)  # Получаем букву столбца
        for row_idx in range(2, row1):

            cell_value = ws.cell(row=row_idx, column=col_idx).value
            cell = ws.cell(row=row_idx, column=col_idx)


            merged = False

            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    merged = True
                    
            if merged:
                continue

            try:
                if len(str(cell_value)) > max_max_len:
                    max_length = max_max_len
                    break

                if len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))

            except:
                pass


        
        adjusted_width = max_length + 2  # Немного добавляем к длине
        
        ws.column_dimensions[column_letter].width = max(
            ws.column_dimensions[column_letter].width, adjusted_width
        )

#        print('>>', column_letter, ws.column_dimensions[column_letter].width)




    # Учитываем ширину объединённых ячеек
    for merged_range in ws.merged_cells.ranges:
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        min_row = merged_range.min_row
        max_row = merged_range.max_row

#        cell.border = border  # Применение границы к ячейке
#        ws.cell(row=min_row, column=min_col).border = thin_border

        merged_cell_value = ws.cell(row=min_row, column=min_col).value

        if merged_cell_value:

            merged_length = len(str(merged_cell_value)) + 2
            if (max_row > min_row):
                words = re.split(r'[,\s\-!/]+', merged_cell_value)
                longest_word = max(words, key=len)
                length_of_longest_word = len(longest_word)

                merged_length = max(merged_length / (max_row-min_row+1), length_of_longest_word) + 2

            # Распределяем длину текста на все столбцы объединенной ячейки
            for col_idx in range(min_col, max_col + 1):
                column_letter = get_column_letter(col_idx)

                ws.column_dimensions[column_letter].width = max(
                    ws.column_dimensions[column_letter].width, merged_length / (max_col - min_col + 1) + 2
                )

    set_default_font(ws)
#    exit(0)


#-------------------------------------------------------------------------------------

def set_default_font(ws):
    # Устанавливаем шрифт по умолчанию для всех ячеек на листе
#    default_font = Font(name="Times New Roman", size=12)  # Например, шрифт Arial размером 12
    return

    for row in ws.iter_rows():
        for cell in row:
            cell.font = cell.font.copy(name="Times New Roman")
#            cell.font = default_font



def adjust_table2(ws, row1, col1, row2, col2):

    for col_idx in range(col1, col2+1):  
        row_m1 = row1
        row_m2 = row1

        old_value = ws.cell(row=row1, column=col_idx).value
        if old_value is None: old_value = ''

        for row_idx in range(row1+1, row2):
            row_m2 = row_idx - 1

            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None: cell_value = ''

            if cell_value == old_value:
                row_m2 = row_idx

            else:
                if row_m1 != row_m2:
                    letter = get_column_letter(col_idx)
                    cell_range = f'{letter}{row_m1}:{letter}{row_m2}'
                    ws.merge_cells(cell_range)
                
                row_m1 = row_idx

            old_value = cell_value    

        row_m2 = row2

        if row_m1 != row_m2:
            letter = get_column_letter(col_idx)
            cell_range = f'{letter}{row_m1}:{letter}{row_m2}'

            ws.merge_cells(cell_range)



def adjust_table2_2(ws, row1, col1, row2, col2):

    row_m1 = row1
    row_m2 = row1

    old_value1 = ws.cell(row=row1, column=col1).value
    old_value2 = ws.cell(row=row1, column=col2).value
    if old_value1 is None: old_value1 = ''
    if old_value2 is None: old_value2 = ''

    for row_idx in range(row1+1, row2):
        row_m2 = row_idx - 1

        cell_value1 = ws.cell(row=row_idx, column=col1).value
        cell_value2 = ws.cell(row=row_idx, column=col2).value

        if cell_value1 is None: cell_value1 = ''
        if cell_value2 is None: cell_value2 = ''

#        print(cell_value1, cell_value2)

        if cell_value1 == old_value1 and cell_value2 == old_value2:
            row_m2 = row_idx

        else:
            if row_m1 != row_m2:
                letter1 = get_column_letter(col1)
                cell_range = f'{letter1}{row_m1}:{letter1}{row_m2}'
                ws.merge_cells(cell_range)
#                print(cell_range)

                letter2 = get_column_letter(col2)
                cell_range = f'{letter2}{row_m1}:{letter2}{row_m2}'
                ws.merge_cells(cell_range)

            row_m1 = row_idx

        old_value1 = cell_value1
        old_value2 = cell_value2

    row_m2 = row2

    if row_m1 != row_m2:
         letter1 = get_column_letter(col1)
         cell_range = f'{letter1}{row_m1}:{letter1}{row_m2}'
         ws.merge_cells(cell_range)

         letter2 = get_column_letter(col2)
         cell_range = f'{letter2}{row_m1}:{letter2}{row_m2}'
         ws.merge_cells(cell_range)



def adjust_table2_3(ws, row1, row2, cols):

    if row1 > row2: 
        return

    row_m1 = row1
    row_m2 = row1

#    print(ws.title, row1, row2)

    old_values = [ws.cell(row=row1, column=col).value for col in cols]

    for row_idx in range(row1+1, row2):
        row_m2 = row_idx - 1

        cell_values = [ws.cell(row=row_idx, column=col).value for col in cols]

        if cell_values == old_values:
            row_m2 = row_idx
        else:
            if row_m1 != row_m2:
                for col in cols:
                    letter = get_column_letter(col)
                    cell_range = f'{letter}{row_m1}:{letter}{row_m2}'
                    ws.merge_cells(cell_range)

            row_m1 = row_idx

        old_values = cell_values

    row_m2 = row2

    if row_m1 != row_m2:
        for col in cols:
            letter = get_column_letter(col)
            cell_range = f'{letter}{row_m1}:{letter}{row_m2}'
            ws.merge_cells(cell_range)



