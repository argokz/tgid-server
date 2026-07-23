import sys
import os
import psycopg2 as pyodbc
import logging
import uuid
import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed, wait

from openpyxl import Workbook

import config
import connect
from psycopg2.extensions import connection as Connection
import views
import excel

import any
import sql
import sql_pass
import ms1
import rs1
import db2

import sort_graph
import f1
import f2_1
import f2_2
import f3
import f4
import f5
import f6
import f7
import f8
import f9
import f10
import f11
import f12
import f13
import f14
import f15

import db2


def read_ms_rs(conn, ms_rs, id):
    q = f'''

select 

--ms.id, ms_rs, 
ms.name,

ue.nomer_uchastka, nu.fio

from (
select ms.id, 'ms' as ms_rs, ms.opisanie_uchastka_ms as name, ms.nomer_uchastka
from uchastok_ms ms
union 
select rs.id, 'rs' as ms_rs, rs.naimenovanie_uchastka_rs, rs.nomer_uchastka
from uchastok_rs rs
) ms

left join uchastki_ekspluatatsii ue ON ue.id = ms.nomer_uchastka
left join nachalniki_uchastkov nu on ue.nachalnik_uchastka=nu.id

where ms.ms_rs='{ms_rs}' and ms.id={id}
'''

    return db2.read_q(conn, q)


def async_do_passport(c, wb, ms_rs, id, fragments, mark_line, mark_pts, mark_node, vals):
    conn = c
    ws1 = wb.create_sheet(title="Ф1.Трубы")
    ws2_1 = wb.create_sheet(title="Ф2_1.Механическое оборудование")
    ws2_2 = wb.create_sheet(title="Ф2_2.Механическое оборудование")
    ws3 = wb.create_sheet(title="Ф3.Каналы")
    ws4 = wb.create_sheet(title="Ф4.Камеры")
    ws5 = wb.create_sheet(title="Ф5.Павильоны")
    ws6 = wb.create_sheet(title="Ф6.Опоры")
    ws7 = wb.create_sheet(title="Ф7.Спец.констр.")
    ws8 = wb.create_sheet(title="Ф8.Изоляция труб")
    ws9 = wb.create_sheet(title="Ф9.Ответств.лицо")
    ws10 = wb.create_sheet(title="Ф10.Ремонт")
    ws11 = wb.create_sheet(title="Ф11.Нарушение")
    ws12 = wb.create_sheet(title="Ф12.Шурфовки")
    ws13 = wb.create_sheet(title="Ф13.Вырезки")
    ws14 = wb.create_sheet(title="Ф14.Опрессовки")
    ws15 = wb.create_sheet(title="Ф15.Осмотр")

    f1.do_passport(conn, ws1, ms_rs, id, fragments, mark_line, mark_pts)
    f2_1.do_passport(conn, ws2_1, ms_rs, id, fragments, mark_line, mark_pts)
    f2_2.do_passport(conn, ws2_2, ms_rs, id, fragments, mark_line, mark_pts)
    f3.do_passport(conn, ws3, ms_rs, id, fragments, mark_line, mark_pts)
    f4.do_passport(conn, ws4, ms_rs, id, fragments, mark_line, mark_pts, mark_node, vals)
    f5.do_passport(conn, ws5, ms_rs, id, fragments, mark_line, mark_pts, mark_node, vals)
    f6.do_passport(conn, ws6, ms_rs, id, fragments, mark_line, mark_pts)
    f7.do_passport(conn, ws7, ms_rs, id, fragments, mark_line, mark_pts)
    f8.do_passport(conn, ws8, ms_rs, id, fragments, mark_line, mark_pts)
    f9.do_passport(conn, ws9, ms_rs, id, fragments, mark_line, mark_pts)
    f10.do_passport(conn, ws10, ms_rs, id, fragments, mark_line, mark_pts)
    f11.do_passport(conn, ws11, ms_rs, id, fragments, mark_line, mark_pts)
    f12.do_passport(conn, ws12, ms_rs, id, fragments, mark_line, mark_pts)
    f13.do_passport(conn, ws13, ms_rs, id, fragments, mark_line, mark_pts)
    f14.do_passport(conn, ws14, ms_rs, id, fragments, mark_line, mark_pts)
    f15.do_passport(conn, ws15, ms_rs, id, fragments, mark_line, mark_pts)
    print("async_do_passport_end")


def passport(**c):
#    try:

    c_without_password = c
    c_without_password.pop('password')

#    print(c)
#    exit(0)

    logging.info(c_without_password)
    
    id = c['id']
    ms_rs = c['ms_rs']
    fragments = c['fragments']
    out_file = c['out_file']

    conn = connect.connect(**c)

    wb = Workbook()

    ws = wb.active

    q = sql_pass.passport(ms_rs, id)
    vals = db2.read_q(conn, q)

    if ms_rs == 'ms':
        ms1.write_ms(ws, vals)
    else:
        rs1.write_rs(ws, vals)

    vals = read_ms_rs(conn, ms_rs, id)

    vals2 = sort_graph.make_graph(conn, fragments, ms_rs, id)
    if vals2 is None:
        return

    mark_line, mark_node, mark_pts = vals2
#    print(mark_pts)
#    print(mark_line)
#    print(mark_node)
    conn.close()

#    ws = wb.create_sheet(title="!!!")
#    f7.do_passport(c, ws, ms_rs, id, fragments, mark_line, mark_pts)
#    exit(0)

    if mark_line == '':
        print("Нет участков")
        exit(1)


    async_do_passport(conn, wb, ms_rs, id, fragments, mark_line, mark_pts, mark_node, vals)

    logging.info(f'Записываем в {out_file}')


    print(os.path.dirname(out_file))
    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    wb.save(out_file)

    print('Выполнено')

#    except Exception as e:
#        print(f"Ошибка export_tgid: {e}")
#    finally:
#        cursor.close()
#        conn.close()
#    conn.close()


def run():
    print('Начали')

    instance_id = str(uuid.uuid4())[:6]

    FORMAT = '%(asctime)s %(levelname)s %(message)s'                    
#    FORMAT = '%(asctime)s %(clientip)-15s %(user)-8s %(message)s'
#    FORMAT = '%(asctime)s %(user)-8s %(message)s'
    FORMAT = f'%(asctime)s {instance_id} %(message)s'                    

    log_date = str(datetime.datetime.today().strftime('%Y_%m_%d'))
    logging.basicConfig(filename=f'{any.argpath_2()}/pas_{log_date}.log',
#                    level=logging.INFO, 
                    level=logging.DEBUG, 
                    filemode='a+',
                    format=FORMAT
                    )

    logging.info('-----------------')
#    logging.info(db)

    logging.info(f'Начали  {instance_id} ')


    path = os.path.splitext(os.path.basename(__file__))[0]
    args = config.init(path, 'Программа для паспортов')

    if not (args.encoding is None): 
#        sys.stdout.reconfigure(encoding='cp866')
        sys.stdout.reconfigure(encoding=args.encoding)

    passport(
          rdbms = args.rdbms,
          server = args.server, 
          user = args.user, 
          password = args.password, 
          db = args.database, 
          port = args.port, 
#          fn = args_fn,
          ms_rs = args.type,
          id = args.id,
          fragments = args.fragments,
          out_file = args.out_file

#          fileID = args.fileID,
#          old_version = args.old_version,
          )

    logging.info(f'Закончили {instance_id} ')


if __name__ == "__main__":

    run()
